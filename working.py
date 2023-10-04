import openpyxl
from datetime import datetime, timedelta

def extract_time(time_str):
    if isinstance(time_str, datetime):
        return time_str
    elif time_str:
        return datetime.strptime(time_str, '%m-%d-%Y %I:%M %p')
    else:
        return None

def extract_hours(timecard_hours_str):
    if timecard_hours_str and isinstance(timecard_hours_str, str):
        try:
            hours, minutes = map(int, timecard_hours_str.split(':'))
            return timedelta(hours=hours, minutes=minutes)
        except ValueError:
            return None
    elif isinstance(timecard_hours_str, timedelta):
        return timecard_hours_str
    else:
        return None




def check_consecutive_days(dates):
    dates.sort()
    for i in range(len(dates) - 1):
        if (dates[i + 1] - dates[i]).days != 1:
            return False
    return True

def check_time_between_shifts(start_time1, end_time1, start_time2, end_time2, min_hours, max_hours):
    time_diff = (start_time2 - end_time1).total_seconds() / 3600
    return min_hours < time_diff < max_hours

def check_single_shift_hours(start_time, end_time, min_hours):
    hours_worked = (end_time - start_time).total_seconds() / 3600
    return hours_worked > min_hours

# Load the Excel file
workbook = openpyxl.load_workbook('Assignment_Timecard.xlsx')

# Assuming you are working with the first sheet
sheet = workbook.active

# Define dictionaries to store relevant information
employee_consecutive_days = {}
employee_shifts_between = {}
employee_long_shifts = {}

# Iterate through rows
for row in sheet.iter_rows(min_row=2, values_only=True):
    employee_name = row[7]
    position_status = row[1]
    start_time = extract_time(row[2])
    end_time = extract_time(row[3])

    if employee_name not in employee_consecutive_days:
        employee_consecutive_days[employee_name] = [start_time]
    else:
        employee_consecutive_days[employee_name].append(start_time)

    if employee_name not in employee_shifts_between:
        employee_shifts_between[employee_name] = []

    if employee_name not in employee_long_shifts:
        employee_long_shifts[employee_name] = []

for employee_name, consecutive_days in employee_consecutive_days.items():
    if check_consecutive_days(consecutive_days):
        print(f"Employee: {employee_name}, Position Status: {position_status} - Worked for 7 consecutive days")

    shifts = employee_shifts_between[employee_name]
    for i in range(len(shifts) - 1):
        start_time1, end_time1 = shifts[i]
        start_time2, end_time2 = shifts[i + 1]

        if start_time1 and end_time1 and start_time2 and end_time2:
            if check_time_between_shifts(start_time1, end_time1, start_time2, end_time2, 1, 10):
                print(f"Employee: {employee_name}, Position Status: {position_status} - Less than 10 hours between shifts")

for employee_name, long_shifts in employee_long_shifts.items():
    for start_time, end_time in long_shifts:
        if start_time and end_time:
            if check_single_shift_hours(start_time, end_time, 14):
                print(f"Employee: {employee_name}, Position Status: {position_status} - Worked for more than 14 hours in a single shift")

# Check for employees who worked more than 14 hours according to the "Timecard Hours (as Time)" column
for row in sheet.iter_rows(min_row=2, values_only=True):
    employee_name = row[7]
    position_status = row[1]
    timecard_hours_str = row[5]  # Assuming the "Timecard Hours (as Time)" column is in the sixth position (index 5)

    timecard_hours = extract_hours(timecard_hours_str)

    if timecard_hours and timecard_hours > timedelta(hours=14):
        print(f"Employee: {employee_name}, Position Status: {position_status} - Worked for more than 14 hours according to Timecard Hours")
